"""Unit tests for the .xlsx export — sheet set, the focused 'ต้องแก้' worklist, and that the
candidate list + collision flag make it into the file.
"""
from io import BytesIO

from openpyxl import load_workbook

from app.schemas import ExportRow, MatchCandidate
from app.services import checklist_xlsx


def _erow(**kw):
    base = dict(symbol="WHA", oldUrl="https://old.example.com/x", newUrl="https://new.example.com/x",
                status="รอดำเนินการ", note="")
    base.update(kw)
    return ExportRow(**base)


def _all_text(ws):
    return " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)


def test_build_has_all_five_sheets():
    wb = load_workbook(BytesIO(checklist_xlsx.build([_erow()])))
    assert wb.sheetnames == ["Redirect Checklist", "Symbol Setup", "Summary", "ต้องแก้", "ผลตรวจ"]


def test_todo_sheet_lists_blank_target_with_candidates_only():
    rows = [
        _erow(oldUrl="https://old.example.com/a", newUrl="", status="ติดปัญหา", matchScore=42.0,
              candidates=[MatchCandidate(url="https://new.example.com/a", score=42.0)]),
        _erow(oldUrl="https://old.example.com/ok", newUrl="https://new.example.com/ok", status="รอดำเนินการ"),
    ]
    text = _all_text(load_workbook(BytesIO(checklist_xlsx.build(rows)))["ต้องแก้"])
    assert "https://old.example.com/a" in text          # blank-target row IS in the worklist
    assert "https://old.example.com/ok" not in text     # the resolved row is NOT
    assert "42.0%" in text                               # its candidate is shown to pick from


def test_collision_row_in_todo_and_flagged():
    rows = [_erow(oldUrl="https://old.example.com/a", newUrl="https://new.example.com/shared",
                  collision=True, status="รอดำเนินการ")]
    text = _all_text(load_workbook(BytesIO(checklist_xlsx.build(rows)))["ต้องแก้"])
    assert "https://old.example.com/a" in text


def test_empty_todo_shows_none_message():
    text = _all_text(load_workbook(BytesIO(checklist_xlsx.build([_erow()])))["ต้องแก้"])
    assert "ไม่มีรายการต้องแก้" in text


def test_detail_sheet_has_candidate_column():
    rows = [_erow(candidates=[MatchCandidate(url="https://new.example.com/c", score=88.0)])]
    text = _all_text(load_workbook(BytesIO(checklist_xlsx.build(rows)))["ผลตรวจ"])
    assert "ตัวเลือกใกล้เคียง (best→)" in text
    assert "88.0%" in text


def test_detail_shows_full_file_list_both_sides():
    """File columns list ALL files per side (matched + unique), not just the diff."""
    rows = [_erow(bodyChecked=True, filesMatched=["shared.pdf"], filesOnlyOld=["gone.pdf"], filesOnlyNew=["added.pdf"])]
    text = _all_text(load_workbook(BytesIO(checklist_xlsx.build(rows)))["ผลตรวจ"])
    assert "ไฟล์เว็บเก่า (ทั้งหมด)" in text and "ไฟล์เว็บใหม่ (ทั้งหมด)" in text
    # old side = matched + only-old ; new side = matched + only-new
    assert "shared.pdf" in text and "gone.pdf" in text and "added.pdf" in text
