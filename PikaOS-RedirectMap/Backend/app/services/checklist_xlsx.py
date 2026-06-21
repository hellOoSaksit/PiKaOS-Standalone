"""Export mapping rows as an .xlsx that matches the central checklist template.

Mirrors `Ref/http_redirect_checklist_5_sites_by_symbol.xlsx` so the export drops straight back
into the team's workflow: three sheets (Redirect Checklist · Symbol Setup · Summary), the same
title rows, the same 7-column header, and the same status dropdown. The tool's 4 statuses are
mapped onto the template's dropdown vocabulary so every exported cell stays valid in Excel.
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..schemas import ExportRow, MappingRow

# tool status → template "ดำเนินการหรือยัง" dropdown value (so the cell is valid in Excel's list)
_STATUS_EXPORT = {
    "รอดำเนินการ": "ยังไม่ดำเนินการ",
    "ดำเนินการแล้ว": "ดำเนินการแล้ว",
    "ติดปัญหา": "ติดปัญหา",
    "ไม่ต้อง Redirect": "ไม่ต้องทำ",
}
# the exact dropdown list from the template (sheet1 E5:E154 data validation)
_DROPDOWN = '"ยังไม่ดำเนินการ,กำลังดำเนินการ,ดำเนินการแล้ว,รอตรวจสอบ,ติดปัญหา,ไม่ต้องทำ"'

_TITLE = "HTTP Redirect Checklist สำหรับเทียบ URL เดิม/ใหม่ และติดตามการทำ Redirect"
_SUBTITLE = "ใช้สำหรับ checklist งาน redirect 5 เว็บไซต์ตาม Symbol: กรอก URL เดิม, URL ใหม่, สถานะ, และ note ของแต่ละรายการ"
_HEADERS = ["No.", "Symbol", "URL เว็บไซต์เดิม", "URL เว็บไซต์ใหม่", "ดำเนินการหรือยัง", "วันที่ดำเนินการ", "Note"]

# styles
_TITLE_FONT = Font(bold=True, size=14)
_SUB_FONT = Font(size=10, italic=True, color="666666")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="4F46E5")
_CENTER = Alignment(horizontal="center", vertical="center")
_WRAP = Alignment(vertical="top", wrap_text=True)
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER


def _sheet_checklist(ws, rows: list[MappingRow]) -> None:
    ws.title = "Redirect Checklist"
    ws.merge_cells("A1:G1")
    ws["A1"] = _TITLE
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:G2")
    ws["A2"] = _SUBTITLE
    ws["A2"].font = _SUB_FONT
    # row 3 blank, row 4 header
    for col, head in enumerate(_HEADERS, 1):
        ws.cell(row=4, column=col, value=head)
    _style_header(ws, 4, len(_HEADERS))

    for i, r in enumerate(rows):
        row = 5 + i
        ws.cell(row=row, column=1, value=i + 1).alignment = _CENTER
        ws.cell(row=row, column=2, value=r.symbol)
        ws.cell(row=row, column=3, value=r.oldUrl)
        ws.cell(row=row, column=4, value=r.newUrl)
        ws.cell(row=row, column=5, value=_STATUS_EXPORT.get(r.status, r.status or ""))
        ws.cell(row=row, column=6, value="")  # วันที่ดำเนินการ — filled by hand when actually done
        ws.cell(row=row, column=7, value=r.note).alignment = _WRAP

    # status dropdown over a generous range (matches the template's E5:E154; extend if more rows)
    last = max(154, 4 + len(rows))
    dv = DataValidation(type="list", formula1=_DROPDOWN, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E5:E{last}")

    widths = [6, 16, 46, 46, 18, 16, 34]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"


def _sheet_symbol_setup(ws, rows: list[MappingRow]) -> None:
    ws.title = "Symbol Setup"
    ws.merge_cells("A1:E1")
    ws["A1"] = "ตั้งค่า Symbol สำหรับเว็บไซต์"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:E2")
    ws["A2"] = "Symbol จริงของแต่ละเว็บไซต์ — ใช้ Symbol เดียวกันในหน้า Redirect Checklist"
    ws["A2"].font = _SUB_FONT
    heads = ["Symbol", "ชื่อเว็บไซต์/บริษัท", "URL หลักเว็บไซต์เดิม", "URL หลักเว็บไซต์ใหม่", "Note"]
    for col, head in enumerate(heads, 1):
        ws.cell(row=4, column=col, value=head)
    _style_header(ws, 4, len(heads))

    # one row per distinct symbol (in first-seen order); main old/new = its first mapping row
    seen: dict[str, MappingRow] = {}
    for r in rows:
        key = r.symbol.strip() or "(no symbol)"
        if key not in seen:
            seen[key] = r
    for i, (sym, r) in enumerate(seen.items()):
        row = 5 + i
        ws.cell(row=row, column=1, value=sym)
        ws.cell(row=row, column=3, value=_origin_of(r.oldUrl))
        ws.cell(row=row, column=4, value=_origin_of(r.newUrl))

    for c, w in enumerate([16, 24, 40, 40, 30], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"


def _sheet_summary(ws, rows: list[MappingRow], as_of: str) -> None:
    ws.title = "Summary"
    ws.merge_cells("A1:H1")
    ws["A1"] = "สรุปสถานะ HTTP Redirect Checklist"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:H2")
    ws["A2"] = f"สร้างเมื่อ {as_of} — ติดตามจำนวน URL ที่ต้อง redirect แยกตาม Symbol"
    ws["A2"].font = _SUB_FONT

    total = len(rows)
    done = sum(1 for r in rows if r.status == "ดำเนินการแล้ว")
    problem = sum(1 for r in rows if r.status == "ติดปัญหา")
    remaining = total - done
    pct = round(done / total * 100, 1) if total else 0

    ws.merge_cells("A4:H4")
    ws["A4"] = "ภาพรวม"
    ws["A4"].font = Font(bold=True)
    over_head = ["รายการทั้งหมด", "ดำเนินการแล้ว", "คงเหลือ", "ติดปัญหา", "% สำเร็จ"]
    for col, head in enumerate(over_head, 1):
        ws.cell(row=5, column=col, value=head)
    _style_header(ws, 5, len(over_head))
    for col, val in enumerate([total, done, remaining, problem, pct], 1):
        ws.cell(row=6, column=col, value=val).alignment = _CENTER

    # per-symbol breakdown
    per_head = ["Symbol", "ชื่อเว็บไซต์/บริษัท", "URL เดิมหลัก", "URL ใหม่หลัก", "รายการทั้งหมด", "ดำเนินการแล้ว", "คงเหลือ", "ติดปัญหา"]
    for col, head in enumerate(per_head, 1):
        ws.cell(row=9, column=col, value=head)
    _style_header(ws, 9, len(per_head))

    order: list[str] = []
    by_sym: dict[str, list[MappingRow]] = {}
    for r in rows:
        key = r.symbol.strip() or "(no symbol)"
        if key not in by_sym:
            by_sym[key] = []
            order.append(key)
        by_sym[key].append(r)
    for i, sym in enumerate(order):
        grp = by_sym[sym]
        s_total = len(grp)
        s_done = sum(1 for r in grp if r.status == "ดำเนินการแล้ว")
        s_problem = sum(1 for r in grp if r.status == "ติดปัญหา")
        row = 10 + i
        ws.cell(row=row, column=1, value=sym)
        ws.cell(row=row, column=3, value=_origin_of(grp[0].oldUrl))
        ws.cell(row=row, column=4, value=_origin_of(grp[0].newUrl))
        ws.cell(row=row, column=5, value=s_total).alignment = _CENTER
        ws.cell(row=row, column=6, value=s_done).alignment = _CENTER
        ws.cell(row=row, column=7, value=s_total - s_done).alignment = _CENTER
        ws.cell(row=row, column=8, value=s_problem).alignment = _CENTER

    for c, w in enumerate([16, 22, 36, 36, 14, 14, 12, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# --- "ผลตรวจ" sheet — mirrors the on-screen table (left block) + the deep findings (right block).
# So a recipient who never opens the web app sees exactly what was on screen, in Excel.
# Left 10 cols = the web table 1:1; right 6 cols = the per-row "▸ detail" (HTTP / redirect / files).

_DETAIL_HEADERS = [
    # --- left block: the on-screen table, column-for-column ---
    "No.", "Symbol", "URL เดิม", "URL ใหม่", "ใกล้เคียง", "ไฟล์", "เนื้อหาเดิม", "เนื้อหาใหม่", "สถานะ", "Note",
    # --- right block: the expandable detail ---
    "HTTP เดิม", "เดิม redirect ไป", "HTTP ใหม่", "ปลายทางจริง (final)", "ไฟล์เว็บเก่า (ทั้งหมด)", "ไฟล์เว็บใหม่ (ทั้งหมด)",
    # --- match decision aids ---
    "ซ้ำปลายทาง", "ตัวเลือกใกล้เคียง (best→)",
]
_BODY_FILL = {
    "error": PatternFill("solid", fgColor="FCE8E6"),   # soft-error page (200-but-broken)
    "thin": PatternFill("solid", fgColor="FEF3C7"),    # H1-only
    "has": PatternFill("solid", fgColor="E7F6EC"),     # real content
    "spa": PatternFill("solid", fgColor="E8F0FE"),     # browser-only (WAF/JS) — body unreadable server-side
}
# status cell tint (keyed by the tool's Thai status) — so a glance down the column shows what's done /
# pending / stuck, same palette as the on-screen chips.
_STATUS_FILL = {
    "ดำเนินการแล้ว": PatternFill("solid", fgColor="E7F6EC"),    # green
    "รอดำเนินการ": PatternFill("solid", fgColor="FEF3C7"),      # amber
    "ติดปัญหา": PatternFill("solid", fgColor="FCE8E6"),         # red
    "ไม่ต้อง Redirect": PatternFill("solid", fgColor="EEEEEE"),  # grey
}
_COLLISION_FILL = PatternFill("solid", fgColor="FFE0B2")        # orange — forced/duplicate target


def _body_text(checked: bool, has_body: bool, thin: bool, error: str, spa: bool = False) -> tuple[str, str]:
    """(label, fill-key) describing one side's body content."""
    if not checked:
        return "—", ""
    if error:
        return f"⚠ หน้า error: {error}", "error"
    if spa:
        return "กันบอต/JS — ตรวจอัตโนมัติไม่ได้", "spa"
    if thin:
        return "มีแค่หัวข้อ H1", "thin"
    if has_body:
        return "มีเนื้อหา", "has"
    return "ว่าง", ""


def _files_text(checked: bool, same, n_old: int, n_new: int, only_old: list, only_new: list) -> str:
    if not checked:
        return "—"
    if same is None:
        return "ไม่มีไฟล์"
    if same:
        return f"เหมือนกัน ({n_old})"
    bits = []
    if only_old:
        bits.append(f"หาย {len(only_old)}")
    if only_new:
        bits.append(f"เพิ่ม {len(only_new)}")
    return f"ต่างกัน เดิม{n_old}/ใหม่{n_new}" + (f" ({', '.join(bits)})" if bits else "")


def _candidates_text(cands) -> str:
    """Multi-line 'score%  url' list of close-match candidates (best first) — what to pick from when
    the auto-match was too weak. Empty when there were none."""
    return "\n".join(f"{c.score}%  {c.url}" for c in (cands or []))


def _needs_fix(r: ExportRow) -> bool:
    """A row a human must act on: stuck, blank target (no strong match), or a forced/duplicate target."""
    return r.status == "ติดปัญหา" or bool(r.collision) or bool(r.oldUrl and not r.newUrl)


def _sheet_detail(ws, rows: list[ExportRow]) -> None:
    ws.title = "ผลตรวจ"
    ncol = len(_DETAIL_HEADERS)
    last = get_column_letter(ncol)
    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = "ผลตรวจทุก URL — เหมือนตารางบนหน้าเว็บ (เปิดดูใน Excel ได้โดยไม่ต้องเข้าเว็บ)"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = ("คอลัมน์ซ้าย (ถึง 'Note') = ตามตารางหน้าเว็บทุกช่อง · คอลัมน์ขวา = รายละเอียดเพิ่ม "
                "(HTTP/redirect/ปลายทางจริง/รายชื่อไฟล์ที่หาย-เพิ่ม)")
    ws["A2"].font = _SUB_FONT
    for col, head in enumerate(_DETAIL_HEADERS, 1):
        ws.cell(row=4, column=col, value=head)
    _style_header(ws, 4, ncol)

    for i, r in enumerate(rows):
        row = 5 + i
        match = "" if r.matchScore is None else f"{r.matchScore}%"
        files = _files_text(r.bodyChecked, r.filesSame, r.oldFileCount, r.newFileCount, r.filesOnlyOld, r.filesOnlyNew)
        old_body, old_key = _body_text(r.bodyChecked, r.oldHasBody, r.oldBodyThin, r.oldError, r.oldSpa)
        new_body, new_key = _body_text(r.bodyChecked, r.newHasBody, r.newBodyThin, r.newError, r.newSpa)
        # full file list per side (matched + that side's unique) so the reader sees what each page
        # actually links, not only the diff — the diff counts stay in the 'ไฟล์' summary column.
        all_old = ", ".join(sorted(set(r.filesMatched) | set(r.filesOnlyOld))) if r.bodyChecked else ""
        all_new = ", ".join(sorted(set(r.filesMatched) | set(r.filesOnlyNew))) if r.bodyChecked else ""
        values = [
            # left block — the web table, column-for-column
            i + 1, r.symbol, r.oldUrl, r.newUrl, match, files, old_body, new_body,
            _STATUS_EXPORT.get(r.status, r.status or ""), r.note,
            # right block — the per-row detail
            (r.oldStatus if r.oldStatus is not None else "—"), r.oldRedirectsTo or "",
            (r.newStatus if r.newStatus is not None else "—"), r.newFinalUrl or "", all_old, all_new,
            # match decision aids
            ("ซ้ำ" if r.collision else ""), _candidates_text(r.candidates),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _BORDER
            if col in (3, 4, 10, 12, 14, 15, 16, 18):  # URLs · Note · redirect/final · file lists · candidates → wrap
                cell.alignment = _WRAP
            elif col in (1, 5, 11, 13, 17):            # No. · match% · HTTP · collision → center
                cell.alignment = _CENTER
            else:
                cell.alignment = Alignment(vertical="top")
        # tint the two body cells by content state so a problem pops on a glance (same colors as the UI)
        if old_key:
            ws.cell(row=row, column=7).fill = _BODY_FILL[old_key]
        if new_key:
            ws.cell(row=row, column=8).fill = _BODY_FILL[new_key]
        # tint the status cell (done/pending/stuck) and flag a forced/duplicate target
        sfill = _STATUS_FILL.get(r.status)
        if sfill:
            ws.cell(row=row, column=9).fill = sfill
        if r.collision:
            ws.cell(row=row, column=17).fill = _COLLISION_FILL

    widths = [6, 14, 44, 44, 10, 22, 14, 14, 16, 40, 9, 30, 9, 30, 40, 40, 11, 50]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "E5"   # keep No./Symbol/URLs on screen while scrolling into the detail cols


# --- "ต้องแก้" sheet — a focused worklist of just the rows a human must act on -----------------
_TODO_HEADERS = ["No.", "Symbol", "URL เดิม", "ปลายทางที่ตั้งไว้", "ใกล้เคียง", "ปัญหา", "ตัวเลือกใกล้เคียง (เลือกอันที่ถูก)"]


def _sheet_todo(ws, rows: list[ExportRow]) -> None:
    """Only the rows needing action (stuck · blank target · forced/duplicate), each with its problem
    note + the close-match candidates to pick from — so a reviewer works one short list, not 300 rows."""
    ws.title = "ต้องแก้"
    todo = [r for r in rows if _needs_fix(r)]
    ws.merge_cells("A1:G1")
    ws["A1"] = f"รายการที่ต้องแก้ ({len(todo)}) — ติดปัญหา · ยังไม่เลือกปลายทาง · ปลายทางซ้ำ"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:G2")
    ws["A2"] = "เลือกปลายทางที่ถูกจากคอลัมน์ 'ตัวเลือกใกล้เคียง' แล้วไปแก้ในชีต Redirect Checklist"
    ws["A2"].font = _SUB_FONT
    for col, head in enumerate(_TODO_HEADERS, 1):
        ws.cell(row=4, column=col, value=head)
    _style_header(ws, 4, len(_TODO_HEADERS))

    if not todo:
        ws.merge_cells("A5:G5")
        ws["A5"] = "✓ ไม่มีรายการต้องแก้"
        ws["A5"].alignment = _CENTER
    for i, r in enumerate(todo):
        row = 5 + i
        match = "" if r.matchScore is None else f"{r.matchScore}%"
        vals = [i + 1, r.symbol, r.oldUrl, r.newUrl or "(ยังไม่เลือก)", match, r.note, _candidates_text(r.candidates)]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = _WRAP if col in (3, 4, 6, 7) else (_CENTER if col in (1, 5) else Alignment(vertical="top"))
        # flag the target cell: orange = forced/duplicate, red = stuck / not yet chosen
        ws.cell(row=row, column=4).fill = _COLLISION_FILL if r.collision else _STATUS_FILL["ติดปัญหา"]

    for c, w in enumerate([6, 14, 46, 46, 10, 44, 52], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"


def _origin_of(url: str) -> str:
    """Scheme://host of a URL (the 'main' URL for a symbol), empty string if not a URL."""
    from urllib.parse import urlsplit
    p = urlsplit((url or "").strip())
    return f"{p.scheme}://{p.netloc}" if p.scheme and p.netloc else (url or "")


def build(rows: list[ExportRow]) -> bytes:
    as_of = datetime.now(timezone.utc).date().isoformat()
    wb = Workbook()
    _sheet_checklist(wb.active, rows)           # template-faithful 7-column checklist
    _sheet_symbol_setup(wb.create_sheet(), rows)
    _sheet_summary(wb.create_sheet(), rows, as_of)
    _sheet_todo(wb.create_sheet(), rows)        # focused worklist: only rows needing action
    _sheet_detail(wb.create_sheet(), rows)      # the full verify findings, one row per URL
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
